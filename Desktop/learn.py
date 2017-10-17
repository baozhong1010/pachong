# -*- coding: utf-8 -*-
import requests
from bs4 import BeautifulSoup
from itertools import product
import types
import openpyxl
from openpyxl import worksheet
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import range_boundaries
#登录部分

root_url = 'http://172.16.203.12/zentao/user-login.html'
index_url = 'http://172.16.203.12/zentao/project-task-206.html'
UA = "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36"

header = {"User-Agent": UA,
           "referer":"http://172.16.203.12/zentao/my/"
           }

r = requests.Session()
f = r.get(root_url,headers = header)

r.cookies = requests.utils.cookiejar_from_dict({
    'zentaosid':'0qorkcc8602i5gg00seenj6kc2'})

r.post(root_url,
       cookies = r.cookies,
       headers = header
       )

#抓取数据部分

f = r.get(index_url,headers = header)

soup = BeautifulSoup(f.content,'lxml')

plans = soup.find_all('tr',class_='text-center')


for plan in plans:
    l = list(plan.stripped_strings)
    status = l[2]
    if status == u'\u8fdb\u884c\u4e2d':
        names = l[1]
        statuss = l[2]
        times = l[3]
        plans = l[-1]
        print names,status,times,plans,'\n'


#names只在上面输出能全部输出，我想再调用names写到表格里发现永远只有一个值，没有其他值了
print names


#填写表格

def patch_worksheet():
    """This monkeypatches Worksheet.merge_cells to remove cell deletion bug
    https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
    Thank you to Sergey Pikhovkin for the fix
    """

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1)
        This is monkeypatched to remove cell deletion bug
        https://bitbucket.org/openpyxl/openpyxl/issues/365/styling-merged-cells-isnt-working
        """
        if not range_string and not all((start_row, start_column, end_row, end_column)):
            msg = "You have to provide a value either for 'coordinate' or for\
            'start_row', 'start_column', 'end_row' *and* 'end_column'"
            raise ValueError(msg)
        elif not range_string:
            range_string = '%s%s:%s%s' % (get_column_letter(start_column),
                                          start_row,
                                          get_column_letter(end_column),
                                          end_row)
        elif ":" not in range_string:
            if COORD_RE.match(range_string):
                return  # Single cell, do nothing
            raise ValueError("Range must be a cell range (e.g. A1:E1)")
        else:
            range_string = range_string.replace('$', '')

        if range_string not in self._merged_cells:
            self._merged_cells.append(range_string)


        # The following is removed by this monkeypatch:

        # min_col, min_row, max_col, max_row = range_boundaries(range_string)
        # rows = range(min_row, max_row+1)
        # cols = range(min_col, max_col+1)
        # cells = product(rows, cols)

        # all but the top-left cell are removed
        #for c in islice(cells, 1, None):
            #if c in self._cells:
                #del self._cells[c]

    # Apply monkey patch
    m = types.MethodType(merge_cells, None, worksheet.Worksheet)
    worksheet.Worksheet.merge_cells = m


patch_worksheet()


wb = load_workbook('OE_Cloud_PMC_2017Q3.xlsx')
ws = wb.active
ws['B22'] = names
ws['E22'] = status + plans
wb.save('OE_Cloud_PMC_2017Q3.xlsx')
